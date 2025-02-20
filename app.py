from flask import Flask, request, render_template, redirect, url_for
from openpyxl import Workbook, load_workbook
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import  MIMEText
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import random

app = Flask(__name__)

USER_TICKETS_FILE = 'xlsx/employee.xlsx'
LOGIN_CREDENTIALS_FILE = 'xlsx/login_credentials.xlsx'
PASSCODE = "984228"

# Initialize login credentials file if it does not exist
if not os.path.exists(LOGIN_CREDENTIALS_FILE):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Employee Name', 'Employee ID', 'Role'])
    sample_data = [
        ['Frank', 'VISTA0001', 'employee'],
        ['Alice', 'VISTA0002', 'employee'],
        ['Bob', 'VISTA0003', 'itsupport'],
        ['Charlie', 'VISTA0004', 'employee'],
        ['Dave', 'VISTA0005', 'itsupport'],
        ['Yogeshwar.s', 'VISTA0001', 'admin'],
        ['Saravanan', 'VISTAMS001', 'superuser'],  # Adding superuser
    ]
    for row in sample_data:
        sheet.append(row)
    workbook.save(LOGIN_CREDENTIALS_FILE)

# Initialize user tickets file if it does not exist
if not os.path.exists(USER_TICKETS_FILE):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Ticket Number', 'Employee Name', 'Employee ID', 'Issue', 'Date', 'Time', 'IT Support', 'Resolution', 'Status'])
    workbook.save(USER_TICKETS_FILE)

def send_email(smtp_server, port, sender, password, recipient, subject, body):
    try:
        msg = MIMEMultipart()
        msg['From'] = sender
        msg['To'] = recipient
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))
        server = smtplib.SMTP(smtp_server, port)
        server.starttls()
        server.login(sender, password)
        text = msg.as_string()
        server.sendmail(sender, recipient, text)
        server.quit()
        print("Email sent successfully!")
    except Exception as e:
        print(f"Failed to send email: {str(e)}")

def generate_unique_ticket_number(sheet):
    existing_numbers = {row[0] for row in sheet.iter_rows(min_row=2, values_only=True)}
    while True:
        ticket_number = random.randint(0, 9999)
        if ticket_number not in existing_numbers:
            return ticket_number

@app.route('/')
def index():
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login():
    employee_name = request.form['ename']
    credential_id = request.form['eid']  # Treat as string
    workbook = load_workbook(LOGIN_CREDENTIALS_FILE)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == employee_name and row[1] == credential_id:
            role = row[2].lower()
            if role == 'employee':
                return redirect(url_for('employee'))
            elif role == 'itsupport':
                return redirect(url_for('itsupport'))
            elif role == 'admin':
                return redirect(url_for('admin'))
            elif role == 'superuser':
                return redirect(url_for('superuser'))
            else:
                return 'Invalid role'
    return render_template('invalid.html')

@app.route('/employee')
def employee():
    return render_template('employee.html')

@app.route('/submit', methods=['POST'])
def submit():
    employee_name = request.form['fname']
    employee_id = request.form['eid']
    issue = request.form['eiss']
    date = request.form['edate']
    time = request.form['etime']
    email = request.form['email']
    
    workbook = load_workbook(USER_TICKETS_FILE)
    sheet = workbook.active
    ticket_number = generate_unique_ticket_number(sheet)
    
    sheet.append([ticket_number, employee_name, employee_id, issue, date, time, '', '', 'Open'])
    workbook.save(USER_TICKETS_FILE)
    
    subject = "Your Details Received"
    body = f"Dear {employee_name},\n\nThank you for informing us of your issue: {issue}. Your ticket number is {ticket_number}. We will review your request and get back to you shortly.\n\nBest regards,\nVISTA Engineering Solutions"
    
    send_email('smtp.gmail.com', 587, 'vistaes17@gmail.com', 'jqelrzqnlpaonqnd', email, subject, body)
    
    return redirect(url_for('thankyou'))

@app.route('/delete_ticket', methods=['POST'])
def delete_ticket():
    ticket_number = int(request.form['ticket_number'])
    passcode = request.form['passcode']
    if passcode == PASSCODE:
        workbook = load_workbook(USER_TICKETS_FILE)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == ticket_number:
                sheet.delete_rows(row[0].row)
                break
        workbook.save(USER_TICKETS_FILE)
        return redirect(url_for('itsupport'))
    else:
        workbook = load_workbook(USER_TICKETS_FILE)
        sheet = workbook.active
        data = list(sheet.values)
        columns = data[0]
        rows = data[1:]
        return render_template('itsupport.html', columns=columns, rows=rows, enumerate=enumerate, invalid_passcode=True)

@app.route('/itsupport')
def itsupport():
    workbook = load_workbook(USER_TICKETS_FILE)
    sheet = workbook.active
    data = list(sheet.values)
    columns = data[0]
    rows = data[1:]
    return render_template('itsupport.html', columns=columns, rows=rows, enumerate=enumerate)

@app.route('/update_ticket', methods=['POST'])
def update_ticket():
    ticket_number = int(request.form['ticket_number'])
    it_support = request.form['it_support']
    resolution = request.form['resolution']
    status = request.form['status']
    workbook = load_workbook(USER_TICKETS_FILE)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == ticket_number:
            row[6].value = it_support
            row[7].value = resolution
            row[8].value = status
            break
    workbook.save(USER_TICKETS_FILE)
    return redirect(url_for('thankyou'))

@app.route('/thankyou')
def thankyou():
    return render_template('thankyou.html')

@app.route('/admin')
def admin():
    workbook = load_workbook(USER_TICKETS_FILE)
    sheet = workbook.active
    data = list(sheet.values)
    columns = data[0]
    rows = data[1:]
    open_count = 0
    closed_count = 0
    for row in rows:
        if row[8] is not None and row[8].lower() == 'open':
            open_count += 1
        elif row[8] is not None and row[8].lower() == 'closed':
            closed_count += 1
    
    labels = ['Open Tickets', 'Closed Tickets']
    sizes = [open_count, closed_count]
    colors = ['#ff9999', '#66b3ff']
    
    fig1, ax1 = plt.subplots()
    ax1.pie(sizes, colors=colors, labels=labels, autopct='%1.1f%%', startangle=90)
    ax1.axis('equal')
    
    pie_chart_filename = 'static/pie_chart.png'
    plt.savefig(pie_chart_filename)
    plt.close(fig1)
    
    return render_template('admin.html', rows=rows, columns=columns, pie_chart=pie_chart_filename)

@app.route('/display')
def display():
    workbook = load_workbook(USER_TICKETS_FILE)
    sheet = workbook.active
    data = list(sheet.values)
    columns = data[0]
    rows = data[1:]
    open_count = 0
    closed_count = 0
    for row in rows:
        if len(row) > 8 and row[8] is not None:
            status = row[8].lower()
            if status == 'open':
                open_count += 1
            elif status == 'closed':
                closed_count += 1
    labels = ['Open Tickets', 'Closed Tickets']
    sizes = [open_count, closed_count]
    colors = ['#ff9999', '#66b3ff']
    fig1, ax1 = plt.subplots()
    ax1.pie(sizes, colors=colors, labels=labels, autopct='%1.1f%%', startangle=90)
    ax1.axis('equal')
    pie_chart_filename = 'static/pie_chart.png'
    plt.savefig(pie_chart_filename)
    return render_template('display.html', rows=rows, columns=columns, pie_chart=pie_chart_filename)

@app.route('/superuser')
def superuser():
    workbook = load_workbook(USER_TICKETS_FILE)
    sheet = workbook.active
    data = list(sheet.values)
    columns = data[0]
    rows = data[1:]
    total_tickets = len(rows)
    return render_template('superuser.html', rows=rows, columns=columns, total_tickets=total_tickets)

@app.route('/itemp')
def itemp():
    workbook = load_workbook(USER_TICKETS_FILE)
    sheet = workbook.active
    data = list(sheet.values)
    columns = data[0]
    rows = data[1:]
    it_team_data = {}
    for row in rows:
        it_support = row[6]
        if it_support:
            if it_support not in it_team_data:
                it_team_data[it_support] = []
            it_team_data[it_support].append(row[:6])
    return render_template('itemp.html', it_team_data=it_team_data)

@app.route('/total')
def total():
    workbook = load_workbook(LOGIN_CREDENTIALS_FILE)
    sheet = workbook.active
    data = list(sheet.values)
    columns = data[0]
    rows = data[1:]
    relevant_data = [(row[0], row[1], row[2]) for row in rows]
    return render_template('total.html', rows=relevant_data, enumerate=enumerate)

@app.route('/delete/<int:index>', methods=['GET', 'POST'])
def delete_employee(index):
    if request.method == 'POST':
        workbook = load_workbook(LOGIN_CREDENTIALS_FILE)
        sheet = workbook.active
        data = list(sheet.values)
        rows = data[1:]
        if index <= len(rows):
            del rows[index - 1]
            sheet.delete_rows(index + 1)
            workbook.save(LOGIN_CREDENTIALS_FILE)
        return redirect('/total')
    else:
        pass

if __name__ == '__main__':
    app.run(host='0.0.0.0', port='5000', debug=True)
